
import pandas as pd
import numpy as np
import io
import re
import openpyxl
from openpyxl.utils import get_column_letter
from config import COMPETENCY_MAP

print("--- [IMPORT] Loading latest report_one_generator.py ---")

# --- Helper function to normalize text for robust matching ---
def normalize_text(text):
    if not isinstance(text, str):
        return ""
    text = re.sub(r'^\d+\.\s*', '', text) # Remove leading numbers and dot
    text = re.sub(r'\s+', '', text)      # Remove all whitespace
    text = re.sub(r'[.,。、？！ー・]', '', text) # Remove common punctuation
    return text.lower()

# --- Main Generator Function ---
# This function is designed to be flexible for different survey periods.
def generate_report_one(df_processed, survey_period):
    import os
    template_path = os.path.join('template', '【その１データ】 RGB意識調査の質問項目と表(職員会議用）.xlsx')

    # Mapping for survey rounds to specific columns in the template
    # Key: Round name (e.g., "第二回"), Value: Dict of grade to column number
    COLUMN_MAPPING = {
        "第一回": {1: 4, 2: 7, 3: 10},  # 4月 in D, G, J (Corrected)
        "第二回": {1: 13, 2: 16, 3: 19}, # 9月 in M, P, S
        "第三回": {1: 22, 2: 25, 3: 28}, # 1月 in V, Y, AB (Corrected)
    }

    # Extract round name like "第二回" from "9月(第二回)"
    # Support both full-width （） and half-width ()
    # Updated Logic for Robust Round Identification
    if "第一回" in survey_period:
        round_name = "第一回"
    elif "第二回" in survey_period:
        round_name = "第二回"
    elif "第三回" in survey_period:
        round_name = "第三回"
    else:
        # Fallback to regex if keyword search fails (for backward compatibility)
        round_name_match = re.search(r'[（\(](.*?)[）\)]', survey_period)
        round_name = round_name_match.group(1) if round_name_match else survey_period
        print(f"[WARNING] Could not detect standard round name (第一回/第二回/第三回) in '{survey_period}'. Using '{round_name}'.")
    
    # Get the column mapping for the current round, or default to second round if not found
    col_map = COLUMN_MAPPING.get(round_name, COLUMN_MAPPING["第二回"])

    # Load the template
    wb = openpyxl.load_workbook(template_path)

    # Only perform grade-based calculations if the '学年' column exists
    if '学年' in df_processed.columns:
        ws = wb['意識調査']

        # Calculate averages from the user's data
        q_averages = {}
        for grade in [1, 2, 3]:
            df_grade = df_processed[df_processed['学年'] == grade]
            q_averages[grade] = {q: df_grade[q].mean() for _, _, qs in COMPETENCY_MAP for q in qs if q in df_grade}

        comp_averages = {}
        for grade in [1, 2, 3]:
            df_grade = df_processed[df_processed['学年'] == grade]
            comp_averages[grade] = {}
            for _, competency, questions in COMPETENCY_MAP:
                valid_qs = [q for q in questions if q in df_processed.columns and not df_processed[q].isnull().all()]
                if valid_qs:
                    avg = df_grade[valid_qs].mean().mean()
                    if pd.isna(avg):
                        avg = 0
                    comp_averages[grade][competency] = avg

        # Pre-compute normalized question map for faster lookup
        # Key: Normalized Text, Value: Original Text
        question_map = {}
        for comp in COMPETENCY_MAP:
            for q in comp[2]:
                normalized_q = normalize_text(q)
                question_map[normalized_q] = q

        # Load the template
        wb = openpyxl.load_workbook(template_path)
        ws = wb.active

        # Iterate through the sheet to find questions and update values
        for row in range(2, ws.max_row + 1):
            # Get question text from the template (Column C)
            cell_val = ws[f'C{row}'].value
            if cell_val:
                # Normalize template question
                q_text = normalize_text(cell_val)
                
                # Find matching question in our data map
                original_q_text = question_map.get(q_text)
                if original_q_text:
                    # Calculate averages per grade
                    avg1 = q_averages.get(1, {}).get(original_q_text, 0)
                    avg2 = q_averages.get(2, {}).get(original_q_text, 0)
                    avg3 = q_averages.get(3, {}).get(original_q_text, 0)
                    
                    ws.cell(row=row, column=col_map[1]).value = avg1 if not pd.isna(avg1) else 0
                    ws.cell(row=row, column=col_map[2]).value = avg2 if not pd.isna(avg2) else 0
                    ws.cell(row=row, column=col_map[3]).value = avg3 if not pd.isna(avg3) else 0
                    
                    # Apply number format
                    for col_idx in col_map.values():
                        if ws.cell(row=row, column=col_idx).value is not None:
                            ws.cell(row=row, column=col_idx).number_format = '0.0'

        # Update the overall competency averages on the right
        # Columns AF/32 (1年), AG/33 (2年), AH/34 (3年)
        for row in range(3, 3 + len(COMPETENCY_MAP)):
            competency_cell = ws[f'AE{row}']
            if competency_cell.value:
                competency_name_in_cell = competency_cell.value.split('(')[0].strip()
                
                for group_name, competency_name, questions in COMPETENCY_MAP:
                    if competency_name_in_cell == competency_name:
                        avg1 = comp_averages.get(1, {}).get(competency_name, 0)
                        avg2 = comp_averages.get(2, {}).get(competency_name, 0)
                        avg3 = comp_averages.get(3, {}).get(competency_name, 0)

                        ws.cell(row=row, column=32).value = avg1 if not pd.isna(avg1) else 0
                        ws.cell(row=row, column=33).value = avg2 if not pd.isna(avg2) else 0
                        ws.cell(row=row, column=34).value = avg3 if not pd.isna(avg3) else 0
                        
                        for col_idx in [32, 33, 34]:
                             if ws.cell(row=row, column=col_idx).value is not None:
                                ws.cell(row=row, column=col_idx).number_format = '0.0'

    # Save to a new in-memory file
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

